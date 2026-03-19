from __future__ import annotations

import io
import threading
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any
from uuid import UUID, uuid4

from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.constants import ExecutionStatus
from app.core.exceptions import AppException
from app.services.provider_service import ProviderRuntimeSelection, ProviderService

settings = get_settings()

MAX_STORED_PROMPT_TESTS = 200


@dataclass(slots=True)
class PromptTestRecord:
    id: UUID
    status: str
    prompt_title: str
    prompt_text: str
    provider_slug: str
    model_slug: str
    file_name: str
    file_size: int
    created_at: datetime
    started_at: datetime | None = None
    finished_at: datetime | None = None
    error_message: str | None = None
    output_text: str | None = None


_STORE_LOCK = threading.Lock()
_PROMPT_TEST_STORE: OrderedDict[str, PromptTestRecord] = OrderedDict()


class PromptTestService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.provider_service = ProviderService(session)

    def create_prompt_test(
        self,
        *,
        prompt_title: str,
        prompt_text: str,
        provider_slug: str,
        model_slug: str,
        file_name: str,
        file_size: int,
    ) -> PromptTestRecord:
        normalized_prompt = str(prompt_text or "").strip()
        normalized_provider = str(provider_slug or "").strip().lower()
        normalized_model = str(model_slug or "").strip().lower()
        normalized_file_name = str(file_name or "").strip()

        if not normalized_prompt:
            raise AppException(
                "Prompt de teste vazio.",
                status_code=400,
                code="prompt_test_empty_prompt",
            )
        if not normalized_provider:
            raise AppException(
                "Provider invalido para teste de prompt.",
                status_code=400,
                code="prompt_test_invalid_provider",
            )
        if not normalized_model:
            raise AppException(
                "Modelo invalido para teste de prompt.",
                status_code=400,
                code="prompt_test_invalid_model",
            )
        if not normalized_file_name:
            raise AppException(
                "Arquivo invalido para teste de prompt.",
                status_code=400,
                code="prompt_test_invalid_file_name",
            )
        if file_size <= 0:
            raise AppException(
                "Arquivo enviado esta vazio.",
                status_code=400,
                code="prompt_test_empty_file",
            )

        # Valida provider/modelo no catalogo operacional antes de enfileirar.
        self.provider_service.resolve_runtime(
            provider_slug=normalized_provider,
            model_slug=normalized_model,
        )

        record = PromptTestRecord(
            id=uuid4(),
            status=ExecutionStatus.QUEUED.value,
            prompt_title=str(prompt_title or "").strip() or "Prompt sem titulo",
            prompt_text=normalized_prompt,
            provider_slug=normalized_provider,
            model_slug=normalized_model,
            file_name=normalized_file_name,
            file_size=int(file_size),
            created_at=datetime.now(timezone.utc),
        )
        self._save_record(record)
        return record

    def get_prompt_test(self, *, prompt_test_id: UUID) -> PromptTestRecord:
        record = self._read_record(prompt_test_id)
        if record is None:
            raise AppException(
                "Teste de prompt nao encontrado.",
                status_code=404,
                code="prompt_test_not_found",
            )
        return record

    def process_prompt_test(
        self,
        *,
        prompt_test_id: UUID,
        file_content: bytes,
    ) -> None:
        record = self._read_record(prompt_test_id)
        if record is None:
            return

        self._update_record(
            prompt_test_id,
            status=ExecutionStatus.PROCESSING.value,
            started_at=datetime.now(timezone.utc),
            finished_at=None,
            error_message=None,
            output_text=None,
        )

        try:
            runtime = self.provider_service.resolve_runtime(
                provider_slug=record.provider_slug,
                model_slug=record.model_slug,
            )
            extracted_content = self._extract_file_content(
                file_name=record.file_name,
                file_content=file_content,
            )
            chunks = self._chunk_content(extracted_content)

            outputs: list[str] = []
            for chunk in chunks:
                prompt_input = self._build_provider_prompt(
                    prompt_text=record.prompt_text,
                    file_content=chunk,
                )
                sanitized_prompt = self._enforce_token_limit(
                    prompt=prompt_input,
                    provider_runtime=runtime,
                )
                result = runtime.client.execute_prompt(
                    prompt=sanitized_prompt,
                    model_name=runtime.model.model_slug,
                    max_tokens=settings.max_tokens,
                    temperature=settings.temperature,
                )
                outputs.append(str(result.output_text or "").strip())

            final_output = "\n\n".join([item for item in outputs if item]).strip()
            self._update_record(
                prompt_test_id,
                status=ExecutionStatus.COMPLETED.value,
                finished_at=datetime.now(timezone.utc),
                output_text=final_output or "(sem retorno textual do modelo)",
                error_message=None,
            )
        except Exception as exc:
            message = self._extract_exception_message(exc)
            self._update_record(
                prompt_test_id,
                status=ExecutionStatus.FAILED.value,
                finished_at=datetime.now(timezone.utc),
                error_message=message,
            )

    def _extract_file_content(self, *, file_name: str, file_content: bytes) -> str:
        lower_name = str(file_name or "").lower()
        if lower_name.endswith(".pdf"):
            return self._extract_pdf_text(file_content)
        if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            return self._extract_xlsx_text(file_content)
        return file_content.decode("utf-8", errors="ignore")

    @staticmethod
    def _extract_pdf_text(content: bytes) -> str:
        try:
            from pypdf import PdfReader
        except Exception:
            return content[:8000].decode("utf-8", errors="ignore")

        try:
            reader = PdfReader(io.BytesIO(content))
            pages: list[str] = []
            for page in reader.pages[:20]:
                pages.append(page.extract_text() or "")
            return "\n".join(pages).strip()
        except Exception:
            return content[:8000].decode("utf-8", errors="ignore")

    @staticmethod
    def _extract_xlsx_text(content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:
            return content[:8000].decode("utf-8", errors="ignore")

        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            rows_text: list[str] = []
            for sheet in workbook.worksheets[:3]:
                for row in sheet.iter_rows(min_row=1, max_row=500, values_only=True):
                    values = [str(value) for value in row if value is not None]
                    if values:
                        rows_text.append(", ".join(values))
            return "\n".join(rows_text).strip()
        except Exception:
            return content[:8000].decode("utf-8", errors="ignore")

    def _chunk_content(self, content: str) -> list[str]:
        normalized = str(content or "").strip()
        if not normalized:
            return ["(arquivo sem conteudo textual)"]

        if len(normalized) > settings.max_input_characters:
            normalized = normalized[: settings.max_input_characters]

        if len(normalized) <= settings.chunk_size_characters:
            return [normalized]

        chunks: list[str] = []
        start = 0
        while start < len(normalized):
            end = min(start + settings.chunk_size_characters, len(normalized))
            chunks.append(normalized[start:end])
            start = end
        return chunks

    @staticmethod
    def _build_provider_prompt(*, prompt_text: str, file_content: str) -> str:
        return (
            "Analise o seguinte arquivo e siga o prompt informado.\n\n"
            f"Conteudo do arquivo:\n{file_content}\n\n"
            f"Prompt:\n{prompt_text}"
        )

    def _enforce_token_limit(
        self,
        *,
        prompt: str,
        provider_runtime: ProviderRuntimeSelection,
    ) -> str:
        max_tokens_allowed = max(int(settings.max_tokens_per_execution), 1)
        current_prompt = str(prompt or "")
        current_tokens = provider_runtime.client.count_tokens(current_prompt)
        if current_tokens <= max_tokens_allowed:
            return current_prompt

        for _ in range(12):
            ratio = max_tokens_allowed / max(current_tokens, 1)
            new_length = max(500, int(len(current_prompt) * ratio * 0.9))
            if new_length >= len(current_prompt):
                new_length = len(current_prompt) - 1
            if new_length <= 0:
                break
            current_prompt = current_prompt[:new_length]
            current_tokens = provider_runtime.client.count_tokens(current_prompt)
            if current_tokens <= max_tokens_allowed:
                return current_prompt

        raise AppException(
            "Prompt excede o limite configurado para teste.",
            status_code=422,
            code="prompt_test_token_limit_exceeded",
            details={"max_tokens_per_execution": settings.max_tokens_per_execution},
        )

    @staticmethod
    def _extract_exception_message(exc: Exception) -> str:
        if isinstance(exc, AppException):
            return str(exc.payload.message or "Falha ao executar teste de prompt.")
        return str(exc or "Falha ao executar teste de prompt.")

    def _save_record(self, record: PromptTestRecord) -> None:
        key = str(record.id)
        with _STORE_LOCK:
            _PROMPT_TEST_STORE[key] = record
            _PROMPT_TEST_STORE.move_to_end(key, last=True)
            while len(_PROMPT_TEST_STORE) > MAX_STORED_PROMPT_TESTS:
                _PROMPT_TEST_STORE.popitem(last=False)

    def _read_record(self, prompt_test_id: UUID) -> PromptTestRecord | None:
        key = str(prompt_test_id)
        with _STORE_LOCK:
            record = _PROMPT_TEST_STORE.get(key)
            if record is None:
                return None
            return PromptTestRecord(**self._record_to_dict(record))

    def _update_record(self, prompt_test_id: UUID, **changes: Any) -> PromptTestRecord | None:
        key = str(prompt_test_id)
        with _STORE_LOCK:
            record = _PROMPT_TEST_STORE.get(key)
            if record is None:
                return None
            updated_payload = self._record_to_dict(record)
            updated_payload.update(changes)
            updated = PromptTestRecord(**updated_payload)
            _PROMPT_TEST_STORE[key] = updated
            _PROMPT_TEST_STORE.move_to_end(key, last=True)
            return updated

    @staticmethod
    def _record_to_dict(record: PromptTestRecord) -> dict[str, Any]:
        return {
            "id": record.id,
            "status": record.status,
            "prompt_title": record.prompt_title,
            "prompt_text": record.prompt_text,
            "provider_slug": record.provider_slug,
            "model_slug": record.model_slug,
            "file_name": record.file_name,
            "file_size": record.file_size,
            "created_at": record.created_at,
            "started_at": record.started_at,
            "finished_at": record.finished_at,
            "error_message": record.error_message,
            "output_text": record.output_text,
        }
