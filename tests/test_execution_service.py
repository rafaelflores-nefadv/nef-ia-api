from datetime import datetime, timezone
import io
from decimal import Decimal
from types import SimpleNamespace
from uuid import UUID, uuid4

from openpyxl import Workbook, load_workbook

from app.core.constants import ExecutionStatus
from app.core.exceptions import AppException
from app.integrations.providers.base import ProviderExecutionResult
from app.models.operational import (
    DjangoAiApiToken,
    DjangoAiApiTokenPermission,
    DjangoAiAutomationExecutionSetting,
    DjangoAiExecutionInputFile,
    DjangoAiQueueJob,
    DjangoAiRequestFile,
)
from app.models.shared import AnalysisExecution
from app.services import execution_service as execution_module
from app.services.execution_engine import (
    EngineExecutionInput,
    ExecutionFileKind,
    build_legacy_tabular_output_contract,
)
from app.services.execution_service import ExecutionService


class FakeSession:
    def __init__(self) -> None:
        self.commit_calls = 0
        self.rollback_calls = 0

    def commit(self) -> None:
        self.commit_calls += 1

    def rollback(self) -> None:
        self.rollback_calls += 1


class FakeRequestFileRepository:
    def __init__(self, request_files: dict[UUID, DjangoAiRequestFile]) -> None:
        self.request_files = request_files

    def get_by_id(self, file_id: UUID) -> DjangoAiRequestFile | None:
        return self.request_files.get(file_id)


class FakeQueueRepository:
    def __init__(self) -> None:
        self.jobs: dict[UUID, DjangoAiQueueJob] = {}
        self.processing_count = 0

    def add(self, job: DjangoAiQueueJob) -> DjangoAiQueueJob:
        if job.id is None:
            job.id = uuid4()
        self.jobs[job.id] = job
        return job

    def get_by_id(self, job_id: UUID) -> DjangoAiQueueJob | None:
        return self.jobs.get(job_id)

    def get_latest_by_execution_id(self, execution_id: UUID) -> DjangoAiQueueJob | None:
        matches = [job for job in self.jobs.values() if job.execution_id == execution_id]
        return matches[-1] if matches else None

    def list_by_execution_id(self, execution_id: UUID) -> list[DjangoAiQueueJob]:
        return [job for job in self.jobs.values() if job.execution_id == execution_id]

    def acquire_for_processing(self, *, queue_job_id: UUID, worker_name: str, started_at) -> bool:  # type: ignore[no-untyped-def]
        job = self.jobs.get(queue_job_id)
        if job is None:
            return False
        if job.job_status not in {ExecutionStatus.QUEUED.value, ExecutionStatus.PENDING.value}:
            return False
        job.job_status = ExecutionStatus.PROCESSING.value
        job.worker_name = worker_name
        job.started_at = started_at
        job.finished_at = None
        job.error_message = None
        return True

    def mark_queued_for_retry(self, *, queue_job_id: UUID, retry_count: int, error_message: str) -> bool:
        job = self.jobs.get(queue_job_id)
        if job is None:
            return False
        job.job_status = ExecutionStatus.QUEUED.value
        job.retry_count = retry_count
        job.error_message = error_message
        job.started_at = None
        job.finished_at = None
        return True

    def count_processing_jobs(self, *, exclude_queue_job_id: UUID | None = None) -> int:
        return self.processing_count


class FakeAuditRepository:
    def __init__(self) -> None:
        self.events: list[object] = []

    def add(self, event: object) -> object:
        self.events.append(event)
        return event


class FakeExecutionInputRepository:
    def __init__(self) -> None:
        self.items: list[object] = []

    def add(self, model):  # type: ignore[no-untyped-def]
        self.items.append(model)
        return model

    def list_by_execution_id(self, execution_id: UUID) -> list[object]:
        return [item for item in self.items if getattr(item, "execution_id", None) == execution_id]

    def get_primary_by_execution_id(self, execution_id: UUID):  # type: ignore[no-untyped-def]
        matches = [
            item
            for item in self.items
            if getattr(item, "execution_id", None) == execution_id and getattr(item, "role", "") == "primary"
        ]
        if not matches:
            return None
        matches.sort(key=lambda item: (int(getattr(item, "order_index", 0)), str(getattr(item, "created_at", ""))))
        return matches[0]


class FakeAutomationExecutionSettingsRepository:
    def __init__(self) -> None:
        self.active_settings: dict[UUID, object] = {}

    def get_active_by_automation_id(self, automation_id: UUID):  # type: ignore[no-untyped-def]
        return self.active_settings.get(automation_id)


class FakeSharedAnalysisRepository:
    def __init__(self, requests: dict[UUID, object]) -> None:
        self.requests = requests

    def get_request_by_id(self, analysis_request_id: UUID):  # type: ignore[no-untyped-def]
        return self.requests.get(analysis_request_id)


class FakeSharedExecutionRepository:
    def __init__(self) -> None:
        self.executions: dict[UUID, AnalysisExecution] = {}

    def create(self, *, analysis_request_id: UUID, status: str) -> AnalysisExecution:
        execution = AnalysisExecution(
            id=uuid4(),
            analysis_request_id=analysis_request_id,
            status=status,
            created_at=datetime.now(timezone.utc),
        )
        self.executions[execution.id] = execution
        return execution

    def get_by_id(self, execution_id: UUID) -> AnalysisExecution | None:
        return self.executions.get(execution_id)

    def list_by_analysis_request_id(self, analysis_request_id: UUID) -> list[AnalysisExecution]:
        return [item for item in self.executions.values() if item.analysis_request_id == analysis_request_id]

    def update_status(self, *, execution_id: UUID, status: str) -> AnalysisExecution | None:
        execution = self.executions.get(execution_id)
        if execution is None:
            return None
        execution.status = status
        return execution


class FakeAutomationRuntimeResolver:
    def __init__(
        self,
        *,
        provider_slug: str = "openai",
        model_slug: str = "gpt-5",
        prompt_text: str = "Prompt oficial",
        automation_slug: str | None = None,
        output_type: str | None = None,
        result_parser: str | None = None,
        result_formatter: str | None = None,
        output_schema: dict | str | None = None,
        debug_enabled: bool = False,
    ) -> None:
        self.provider_slug = provider_slug
        self.model_slug = model_slug
        self.prompt_text = prompt_text
        self.automation_slug = automation_slug
        self.output_type = output_type
        self.result_parser = result_parser
        self.result_formatter = result_formatter
        self.output_schema = output_schema
        self.debug_enabled = debug_enabled
        self.resolve_calls: list[UUID] = []
        self.resolve_require_prompt_calls: list[bool] = []

    def resolve(self, automation_id, *, require_prompt: bool = True):  # type: ignore[no-untyped-def]
        self.resolve_calls.append(automation_id)
        self.resolve_require_prompt_calls.append(require_prompt)
        return SimpleNamespace(
            automation_id=automation_id,
            prompt_text=self.prompt_text,
            prompt_version=3,
            provider_slug=self.provider_slug,
            model_slug=self.model_slug,
            automation_slug=self.automation_slug,
            output_type=self.output_type,
            result_parser=self.result_parser,
            result_formatter=self.result_formatter,
            output_schema=self.output_schema,
            debug_enabled=self.debug_enabled,
        )


class FakeFileService:
    def __init__(self) -> None:
        self.calls: list[dict] = []

    def register_generated_execution_file(self, **kwargs):  # type: ignore[no-untyped-def]
        self.calls.append(kwargs)
        return SimpleNamespace(id=uuid4())


def _legacy_tabular_runtime_kwargs() -> dict[str, object]:
    contract = build_legacy_tabular_output_contract()
    schema = contract.output_schema
    payload: dict[str, object] = {
        "output_type": contract.output_type.value,
        "result_parser": contract.parser_strategy.value,
        "result_formatter": contract.formatter_strategy.value,
        "output_schema": {
            "columns": list(schema.columns),
            "structured_output_aliases": {
                field_name: list(aliases)
                for field_name, aliases in schema.structured_output_aliases.items()
            },
            "prompt_field_columns": dict(schema.prompt_field_columns),
            "prompt_field_aliases": {
                field_name: list(aliases)
                for field_name, aliases in schema.prompt_field_aliases.items()
            },
            "prompt_placeholders": dict(schema.prompt_placeholders),
            "row_origin_column": schema.row_origin_column,
            "status_column": schema.status_column,
            "error_column": schema.error_column,
            "include_input_columns": schema.include_input_columns,
            "input_collision_prefix": schema.input_collision_prefix,
            "worksheet_name": schema.worksheet_name,
            "file_name_template": schema.file_name_template,
            "mime_type": schema.mime_type,
        },
    }
    output_schema = payload["output_schema"]
    if isinstance(output_schema, dict) and schema.ai_output_columns:
        output_schema["ai_output_columns"] = list(schema.ai_output_columns)
    return payload


class FakeProviderClient:
    def __init__(
        self,
        *,
        modes: list[str] | None = None,
        input_tokens: int = 150,
        output_tokens: int = 75,
        estimated_cost: Decimal = Decimal("0.022500"),
        output_text: str = (
            "Classificacao da planilha: Classe A\n"
            "Classificacao correta: Classe B\n"
            "Veredito: Divergente\n"
            "Motivo: Fundamentacao teste\n"
            "Trecho determinante: Trecho teste"
        ),
    ) -> None:
        self.modes = modes or ["success"]
        self.input_tokens = input_tokens
        self.output_tokens = output_tokens
        self.estimated_cost = estimated_cost
        self.output_text = output_text
        self.execute_calls: list[dict] = []

    def execute_prompt(self, **kwargs):  # type: ignore[no-untyped-def]
        self.execute_calls.append(kwargs)
        mode = self.modes.pop(0) if self.modes else "success"
        if mode == "timeout":
            raise AppException("Provider request timed out.", status_code=504, code="provider_timeout")
        if mode == "network":
            raise AppException("Network error.", status_code=502, code="provider_network_error")
        if mode == "unsupported_parameter":
            raise AppException(
                "Provider HTTP 400: Unsupported parameter: 'max_tokens' is not supported with this model.",
                status_code=502,
                code="provider_http_error",
                details={
                    "status_code": 400,
                    "http_status_code": 400,
                    "provider_error_message": "Unsupported parameter: 'max_tokens' is not supported with this model.",
                    "provider_error_type": "invalid_request_error",
                    "provider_error_code": "unsupported_parameter",
                    "provider_error_classification": "provider_unsupported_parameter",
                },
            )
        if mode == "rate_limit":
            raise AppException(
                "Provider HTTP 429: Rate limit exceeded.",
                status_code=502,
                code="provider_http_error",
                details={
                    "status_code": 429,
                    "http_status_code": 429,
                    "provider_error_message": "Rate limit exceeded.",
                    "provider_error_type": "rate_limit_error",
                    "provider_error_code": "rate_limit_exceeded",
                    "provider_error_classification": "provider_rate_limit",
                    "response_headers_relevantes": {"retry-after": "2"},
                },
            )
        if mode == "logic_error":
            raise AppException("Invalid input.", status_code=422, code="invalid_input")
        return ProviderExecutionResult(
            output_text=self.output_text,
            input_tokens=self.input_tokens,
            output_tokens=self.output_tokens,
            raw_response={"id": "resp_test"},
        )

    def count_tokens(self, text: str) -> int:
        return max(1, len(text) // 4)

    def estimate_cost(self, **kwargs):  # type: ignore[no-untyped-def]
        return self.estimated_cost


class FakeProviderService:
    def __init__(self, client: FakeProviderClient, *, resolve_error: AppException | None = None) -> None:
        self.client = client
        self.resolve_error = resolve_error
        self.resolve_calls: list[tuple[str, str]] = []
        self.provider_id = uuid4()
        self.model_id = uuid4()

    def resolve_runtime(
        self,
        *,
        provider_slug: str,
        model_slug: str,
        credential_id=None,
    ):  # type: ignore[no-untyped-def]
        self.resolve_calls.append((provider_slug, model_slug))
        if self.resolve_error is not None:
            raise self.resolve_error
        return SimpleNamespace(
            provider=SimpleNamespace(id=self.provider_id, slug=provider_slug),
            model=SimpleNamespace(
                id=self.model_id,
                model_slug=model_slug,
                cost_input_per_1k_tokens=Decimal("0.150000"),
                cost_output_per_1k_tokens=Decimal("0.600000"),
            ),
            credential=SimpleNamespace(id=uuid4()),
            client=self.client,
        )


class FakeUsageService:
    def __init__(self) -> None:
        self.calls: list[dict] = []

    def register_usage(self, **kwargs):  # type: ignore[no-untyped-def]
        self.calls.append(kwargs)
        return SimpleNamespace(id=uuid4())


def _build_permission(automation_id: UUID) -> DjangoAiApiTokenPermission:
    return DjangoAiApiTokenPermission(
        token_id=uuid4(),
        automation_id=automation_id,
        provider_id=None,
        allow_execution=True,
        allow_file_upload=False,
    )


def _build_api_token() -> DjangoAiApiToken:
    return DjangoAiApiToken(
        id=uuid4(),
        name="exec-token",
        token_hash="hash",
        is_active=True,
        expires_at=None,
        created_by_user_id=uuid4(),
    )


def _build_request_file(
    analysis_request_id: UUID,
    *,
    file_name: str = "input.pdf",
    file_path: str | None = None,
    mime_type: str = "text/csv",
) -> DjangoAiRequestFile:
    return DjangoAiRequestFile(
        id=uuid4(),
        analysis_request_id=analysis_request_id,
        file_name=file_name,
        file_path=file_path or f"requests/test/{uuid4()}_{file_name}",
        file_size=25,
        mime_type=mime_type,
        checksum="checksum",
        uploaded_at=datetime.now(timezone.utc),
    )


def _seed_execution_and_job(
    *,
    shared_exec_repo: FakeSharedExecutionRepository,
    queue_repo: FakeQueueRepository,
    analysis_request_id: UUID,
    request_file_id: UUID,
    status: ExecutionStatus = ExecutionStatus.QUEUED,
    retry_count: int = 0,
) -> tuple[AnalysisExecution, DjangoAiQueueJob]:
    execution = shared_exec_repo.create(analysis_request_id=analysis_request_id, status=status.value)
    queue_job = DjangoAiQueueJob(
        id=uuid4(),
        execution_id=execution.id,
        request_file_id=request_file_id,
        job_status=ExecutionStatus.QUEUED.value,
        retry_count=retry_count,
    )
    queue_repo.add(queue_job)
    return execution, queue_job


def _build_service(
    analysis_request_id: UUID,
    automation_id: UUID,
    request_file: DjangoAiRequestFile,
) -> tuple[ExecutionService, FakeQueueRepository, FakeSharedExecutionRepository]:
    operational_session = FakeSession()
    shared_session = FakeSession()
    service = ExecutionService(operational_session=operational_session, shared_session=shared_session)  # type: ignore[arg-type]
    queue_repo = FakeQueueRepository()
    shared_exec_repo = FakeSharedExecutionRepository()
    service.request_files = FakeRequestFileRepository({request_file.id: request_file})  # type: ignore[assignment]
    service.execution_inputs = FakeExecutionInputRepository()  # type: ignore[assignment]
    service.execution_profile_settings = FakeAutomationExecutionSettingsRepository()  # type: ignore[assignment]
    service.queue_jobs = queue_repo  # type: ignore[assignment]
    service.audit_logs = FakeAuditRepository()  # type: ignore[assignment]
    service.shared_analysis = FakeSharedAnalysisRepository(
        {analysis_request_id: SimpleNamespace(id=analysis_request_id, automation_id=automation_id)}
    )  # type: ignore[assignment]
    service.shared_executions = shared_exec_repo  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver()  # type: ignore[assignment]
    service.file_service = FakeFileService()  # type: ignore[assignment]
    service.usage_service = FakeUsageService()  # type: ignore[assignment]
    service.execution_explanations = SimpleNamespace(
        items={},
        upsert_simple_explanation=lambda execution_id, simple_explanation: SimpleNamespace(
            execution_id=execution_id,
            simple_explanation=simple_explanation,
        ),
    )  # type: ignore[assignment]
    return service, queue_repo, shared_exec_repo


def test_create_execution_creates_queue_job_and_dispatches(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    permissions = [_build_permission(automation_id)]
    dispatched: list[tuple[UUID, UUID, int | None]] = []

    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: dispatched.append((execution_id, queue_job_id, delay_ms)),
    )

    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    result = service.create_execution(
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
        api_token=_build_api_token(),
        token_permissions=permissions,
    )

    assert result.status == ExecutionStatus.QUEUED
    assert result.execution_id in shared_exec_repo.executions
    assert result.queue_job_id in queue_repo.jobs
    assert len(dispatched) == 1
    assert dispatched[0][2] is None


def test_create_execution_supports_request_file_ids_and_keeps_primary(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    primary_file = _build_request_file(analysis_request_id, file_name="primary.csv")
    context_file = _build_request_file(analysis_request_id, file_name="context.csv")
    permissions = [_build_permission(automation_id)]
    dispatched: list[tuple[UUID, UUID, int | None]] = []

    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: dispatched.append((execution_id, queue_job_id, delay_ms)),
    )

    service, queue_repo, _ = _build_service(analysis_request_id, automation_id, primary_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            primary_file.id: primary_file,
            context_file.id: context_file,
        }
    )

    result = service.create_execution(
        analysis_request_id=analysis_request_id,
        request_file_id=None,
        request_file_ids=[primary_file.id, context_file.id],
        api_token=_build_api_token(),
        token_permissions=permissions,
    )

    queue_job = queue_repo.jobs[result.queue_job_id]
    assert queue_job.request_file_id == primary_file.id
    assert len(service.execution_inputs.items) == 2  # type: ignore[attr-defined]
    assert dispatched


def test_create_execution_supports_input_files_roles(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    first_file = _build_request_file(analysis_request_id, file_name="f1.csv")
    second_file = _build_request_file(analysis_request_id, file_name="f2.csv")
    permissions = [_build_permission(automation_id)]

    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: None,
    )

    service, queue_repo, _ = _build_service(analysis_request_id, automation_id, first_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            first_file.id: first_file,
            second_file.id: second_file,
        }
    )

    result = service.create_execution(
        analysis_request_id=analysis_request_id,
        request_file_id=None,
        input_files=[
            {"request_file_id": second_file.id, "role": "context", "order_index": 1},
            {"request_file_id": first_file.id, "role": "primary", "order_index": 0},
        ],
        api_token=_build_api_token(),
        token_permissions=permissions,
    )

    queue_job = queue_repo.jobs[result.queue_job_id]
    assert queue_job.request_file_id == first_file.id
    roles = {item.request_file_id: item.role for item in service.execution_inputs.items}  # type: ignore[attr-defined]
    assert roles[first_file.id] == "primary"
    assert roles[second_file.id] == "context"


def test_execution_with_prompt_override_skips_official_prompt_requirement(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    runtime_resolver = FakeAutomationRuntimeResolver(prompt_text="")
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.runtime_resolver = runtime_resolver  # type: ignore[assignment]
    service.provider_service = provider_service  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    queue_job.prompt_override_text = "Prompt override para execucao"

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert runtime_resolver.resolve_calls == [automation_id]
    assert runtime_resolver.resolve_require_prompt_calls == [False]
    assert provider_service.client.execute_calls
    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "Prompt override para execucao" in sent_prompt


def test_list_execution_inputs_returns_legacy_queue_fallback() -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="legacy.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    queue_job.job_status = ExecutionStatus.COMPLETED.value

    inputs = service.list_execution_inputs(
        execution_id=execution.id,
        token_permissions=[_build_permission(automation_id)],
    )
    assert len(inputs) == 1
    assert inputs[0].request_file_id == request_file.id
    assert inputs[0].role == "primary"
    assert inputs[0].source == "legacy_queue_job"


def test_execution_uses_exact_provider_and_model_from_shared(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert provider_service.resolve_calls == [("openai", "gpt-5")]
    assert provider_service.client.execute_calls
    assert provider_service.client.execute_calls[0]["model_name"] == "gpt-5"


def test_fails_when_provider_is_inactive(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(),
        resolve_error=AppException("Provider inactive.", status_code=422, code="provider_inactive"),
    )

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "inactive" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_fails_when_model_is_inactive(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(),
        resolve_error=AppException("Model inactive.", status_code=422, code="provider_model_inactive"),
    )

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert "model inactive" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_fails_when_model_does_not_belong_to_provider(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="claude-sonnet")  # type: ignore[assignment]
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(),
        resolve_error=AppException("Model mismatch.", status_code=422, code="provider_model_mismatch"),
    )

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert "mismatch" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_fails_when_no_active_credential(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(),
        resolve_error=AppException("No active credential found.", status_code=422, code="provider_credential_not_found"),
    )

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert "credential" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_no_automatic_fallback_when_provider_times_out(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    provider_service = FakeProviderService(FakeProviderClient(modes=["timeout"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    monkeypatch.setattr(execution_module.settings, "max_retries", 3)
    monkeypatch.setattr(execution_module.settings, "retry_backoff_seconds", 2)
    dispatched: list[int | None] = []
    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: dispatched.append(delay_ms),
    )

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.QUEUED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.QUEUED.value
    assert queue_repo.jobs[queue_job.id].retry_count == 1
    assert provider_service.resolve_calls == [("openai", "gpt-5")]
    assert len(provider_service.client.execute_calls) == 1
    assert dispatched and dispatched[0] == 2000


def test_retry_keeps_same_provider_and_model(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.runtime_resolver = FakeAutomationRuntimeResolver(provider_slug="openai", model_slug="gpt-5")  # type: ignore[assignment]
    provider_service = FakeProviderService(FakeProviderClient(modes=["timeout", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    monkeypatch.setattr(execution_module.settings, "max_retries", 3)
    monkeypatch.setattr(execution_module.settings, "retry_backoff_seconds", 1)
    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: None,
    )

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.COMPLETED.value
    assert provider_service.resolve_calls == [("openai", "gpt-5"), ("openai", "gpt-5")]
    assert len(provider_service.client.execute_calls) == 2


def test_idempotency_skips_when_execution_already_completed(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    execution = shared_exec_repo.create(analysis_request_id=analysis_request_id, status=ExecutionStatus.COMPLETED.value)
    queue_job = DjangoAiQueueJob(
        id=uuid4(),
        execution_id=execution.id,
        request_file_id=request_file.id,
        job_status=ExecutionStatus.QUEUED.value,
        retry_count=0,
    )
    queue_repo.add(queue_job)

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert len(provider_service.client.execute_calls) == 0
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.QUEUED.value


def test_cost_limit_aborts_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    expensive_provider = FakeProviderService(
        FakeProviderClient(modes=["success"], estimated_cost=Decimal("999.000000"))
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = expensive_provider  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo arquivo")
    monkeypatch.setattr(execution_module.settings, "max_cost_per_execution", 0.10)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "estimated cost limit" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_token_limit_truncates_prompt(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    provider = FakeProviderService(FakeProviderClient(modes=["success"]))
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = provider  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "A" * 4000)
    monkeypatch.setattr(execution_module.settings, "max_tokens_per_execution", 600)
    monkeypatch.setattr(execution_module.settings, "chunk_size_characters", 4000)
    monkeypatch.setattr(execution_module.settings, "max_input_characters", 4000)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert provider.client.execute_calls
    prompt_sent = provider.client.execute_calls[0]["prompt"]
    assert provider.client.count_tokens(prompt_sent) <= 600


def test_concurrency_limit_schedules_retry(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id)
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    queue_repo.processing_count = 10

    monkeypatch.setattr(execution_module.settings, "max_concurrent_executions", 2)
    monkeypatch.setattr(execution_module.settings, "max_retries", 3)
    monkeypatch.setattr(execution_module.settings, "retry_backoff", 1)
    monkeypatch.setattr(execution_module.settings, "retry_backoff_seconds", 1)
    dispatched: list[int | None] = []
    monkeypatch.setattr(
        "app.services.execution_service.enqueue_execution_job",
        lambda *, execution_id, queue_job_id, correlation_id=None, delay_ms=None: dispatched.append(delay_ms),
    )

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.QUEUED.value
    assert queue_repo.jobs[queue_job.id].retry_count == 1
    assert dispatched and dispatched[0] == 1000


def test_tabular_csv_processes_each_row_and_generates_xlsx(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=(
            "Conteudo: {{CONTEUDO}}\n"
            "Prazo: {{PRAZO_AGENDADO}}\n"
            "Valor: {{VALOR_DA_CAUSA}}\n"
            "Acao: {{TIPO_DE_ACAO}}"
        ),
        **_legacy_tabular_runtime_kwargs(),
    )

    csv_payload = (
        "processo,conteudo,prazo agendado,valor da causa,tipo de acao\n"
        "P1,Conteudo linha 1,2026-01-10,1000,Acao A\n"
        "P2,Conteudo linha 2,2026-01-11,2000,Acao B\n"
        "P3,Conteudo linha 3,2026-01-12,3000,Acao C\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 3
    assert "Conteudo linha 1" in provider_service.client.execute_calls[0]["prompt"]

    assert service.file_service.calls
    generated = service.file_service.calls[0]
    assert generated["file_type"] == "output"
    assert generated["file_name"].endswith("_resultado.xlsx")
    assert generated["mime_type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(io.BytesIO(generated["content"]))
    sheet = workbook.active
    header = [str(item or "") for item in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    assert "classificacao_correta" in header
    assert "veredito" in header


def test_tabular_csv_debug_mode_registers_output_and_debug_files(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text=(
                '{"classificacao_da_planilha":"A",'
                '"classificacao_correta":"B",'
                '"veredito":"Divergente",'
                '"motivo":"Regra",'
                '"trecho_determinante":"Trecho X"}'
            ),
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        debug_enabled=True,
        **_legacy_tabular_runtime_kwargs(),
    )
    csv_payload = (
        "conteudo,prazo agendado\n"
        "Conteudo linha 1,2026-01-10\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(service.file_service.calls) == 2

    output_file = next(item for item in service.file_service.calls if item["file_type"] == "output")
    debug_file = next(item for item in service.file_service.calls if item["file_type"] == "debug")
    assert output_file["file_name"].endswith("_resultado.xlsx")
    assert debug_file["file_name"].startswith("debug_")
    assert debug_file["mime_type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    debug_wb = load_workbook(io.BytesIO(debug_file["content"]))
    debug_sheet = debug_wb.active
    debug_header = [str(item or "") for item in next(debug_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    assert "prompt_template" in debug_header
    assert "prompt_final" in debug_header
    assert "response_raw_text" in debug_header
    assert "json_payload_cleaned" in debug_header
    assert "json_payload_parsed" in debug_header
    assert "projected_output_row" in debug_header
    assert "provider_name" in debug_header
    assert "model_name" in debug_header
    assert "resolved_model_identifier" in debug_header
    assert "request_url" in debug_header
    assert "request_method" in debug_header
    assert "request_timeout_seconds" in debug_header
    assert "api_family_resolved" in debug_header
    assert "request_profile_resolved" in debug_header
    assert "token_limit_param_used" in debug_header
    assert "client_request_id" in debug_header
    assert "request_payload_sanitized" in debug_header
    assert "started_at" in debug_header
    assert "finished_at" in debug_header
    assert "duration_ms" in debug_header
    assert "http_status_code" in debug_header
    assert "provider_error_message" in debug_header
    assert "provider_request_id" in debug_header
    assert "stage_of_failure" in debug_header
    assert "error_type" in debug_header
    debug_rows = list(debug_sheet.iter_rows(min_row=2, values_only=True))
    assert any(str(row[debug_header.index("json_payload_cleaned")] or "").strip() for row in debug_rows)
    assert any(str(row[debug_header.index("json_payload_parsed")] or "").strip() for row in debug_rows)
    assert any(str(row[debug_header.index("request_payload_sanitized")] or "").strip() for row in debug_rows)
    # 1 linha de metadados + 1 linha processada.
    assert debug_sheet.max_row >= 3


def test_tabular_csv_debug_mode_captures_empty_provider_response_context(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text="",
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        debug_enabled=True,
        **_legacy_tabular_runtime_kwargs(),
    )
    csv_payload = "conteudo\nLinha vazia de retorno\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(service.file_service.calls) == 2

    output_file = next(item for item in service.file_service.calls if item["file_type"] == "output")
    output_wb = load_workbook(io.BytesIO(output_file["content"]))
    output_sheet = output_wb.active
    output_header = [str(item or "") for item in next(output_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    output_row = list(output_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert str(output_row[output_header.index("status")] or "") == "erro"
    assert "Provider returned empty body" in str(output_row[output_header.index("erro")] or "")

    debug_file = next(item for item in service.file_service.calls if item["file_type"] == "debug")
    debug_wb = load_workbook(io.BytesIO(debug_file["content"]))
    debug_sheet = debug_wb.active
    debug_header = [str(item or "") for item in next(debug_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    debug_rows = list(debug_sheet.iter_rows(min_row=2, values_only=True))
    data_row = next(row for row in debug_rows if int(row[debug_header.index("row_index")] or 0) > 0)

    assert str(data_row[debug_header.index("stage_of_failure")] or "") == "provider_response_validation"
    assert str(data_row[debug_header.index("error_type")] or "") == "provider_empty_response"
    assert "Provider returned empty body" in str(data_row[debug_header.index("errors")] or "")
    assert str(data_row[debug_header.index("request_payload_sanitized")] or "").strip()
    assert str(data_row[debug_header.index("retry_count")] or "") == "0"


def test_tabular_debug_classifies_openai_unsupported_parameter_without_model_mismatch(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["unsupported_parameter"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        debug_enabled=True,
        provider_slug="openai",
        model_slug="gpt-5-mini",
        **_legacy_tabular_runtime_kwargs(),
    )
    csv_payload = "conteudo\nLinha teste\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    debug_file = next(item for item in service.file_service.calls if item["file_type"] == "debug")
    debug_wb = load_workbook(io.BytesIO(debug_file["content"]))
    debug_sheet = debug_wb.active
    debug_header = [str(item or "") for item in next(debug_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    debug_rows = list(debug_sheet.iter_rows(min_row=2, values_only=True))
    data_row = next(row for row in debug_rows if int(row[debug_header.index("row_index")] or 0) > 0)

    assert str(data_row[debug_header.index("error_type")] or "") == "provider_unsupported_parameter"
    assert str(data_row[debug_header.index("error_type")] or "") != "provider_unsupported_model"
    assert "unsupported parameter" in str(data_row[debug_header.index("errors")] or "").lower()


def test_tabular_csv_row_error_does_not_abort_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success", "logic_error", "success"])
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}"
        ,
        **_legacy_tabular_runtime_kwargs(),
    )

    csv_payload = (
        "conteudo,prazo agendado\n"
        "Conteudo linha 1,2026-01-10\n"
        "Conteudo linha 2,2026-01-11\n"
        "Conteudo linha 3,2026-01-12\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 3
    assert len(service.usage_service.calls) == 2

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    sheet = workbook.active
    header = [str(item or "") for item in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    status_index = header.index("status")
    error_index = header.index("erro")
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    statuses = [str(row[status_index] or "") for row in rows]
    errors = [str(row[error_index] or "") for row in rows]
    assert statuses == ["ok", "erro", "ok"]
    assert "Invalid input." in errors[1]


def test_tabular_rejects_legacy_xls_extension(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.xls")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: b"legacy")

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].error_message == "Legacy .xls files are not supported. Convert the spreadsheet to .xlsx."


def test_tabular_json_output_is_parsed(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text=(
                '{"classificacao_da_planilha":"A","classificacao_correta":"B",'
                '"veredito":"Divergente","motivo":"Regra","trecho_determinante":"Trecho X"}'
            ),
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}"
        ,
        **_legacy_tabular_runtime_kwargs(),
    )

    csv_payload = "conteudo\nLinha unica\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    sheet = workbook.active
    header = [str(item or "") for item in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    verdict_index = header.index("veredito")
    motivo_index = header.index("motivo")
    assert row[verdict_index] == "Divergente"
    assert row[motivo_index] == "Regra"


def test_tabular_automation_contract_customizes_output_schema(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text="Classificacao custom: Divergente\nObservacao: Regra personalizada",
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        automation_slug="audit_contract_test",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": [
                "linha_origem",
                "conteudo",
                "classificacao_custom",
                "observacao",
                "status",
                "erro",
            ],
            "structured_output_aliases": {
                "classificacao_custom": ["classificacao custom", "classificacao_custom"],
                "observacao": ["observacao"],
            },
            "prompt_field_columns": {"conteudo": "conteudo"},
            "worksheet_name": "auditoria",
            "file_name_template": "audit_{execution_id}.xlsx",
        },
    )

    csv_payload = "conteudo\nLinha unica\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    assert generated["file_name"].startswith("audit_")
    workbook = load_workbook(io.BytesIO(generated["content"]))
    sheet = workbook.active
    assert sheet.title == "auditoria"
    header = [str(item or "") for item in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert "classificacao_custom" in header
    assert "observacao" in header
    assert row[header.index("classificacao_custom")] == "Divergente"
    assert row[header.index("observacao")] == "Regra personalizada"


def test_tabular_explicit_schema_projects_exact_columns_and_maps_input(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text=(
                "descricao: Nao deve sobrescrever\n"
                "categoria: Trabalhista\n"
                "pensamento: Texto de analise\n"
                "prazo: 22/03/2026\n"
                "compromissoAnalista: Revisar e protocolar\n"
                "necessitaRevisao: nao\n"
                "resumo_do_andamento: Andamento resumido"
            ),
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    expected_columns = [
        "numero_processo",
        "id_processo",
        "descricao",
        "celula",
        "valor_da_causa",
        "tipo_de_acao",
        "marcacao",
        "responsavel",
        "categoria",
        "pensamento",
        "reclassificacao",
        "prazo",
        "compromissoAnalista",
        "necessitaRevisao",
        "resumo_do_andamento",
    ]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Descricao: {{DESCRICAO}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": expected_columns,
            "structured_output_aliases": {
                "categoria": ["categoria"],
                "pensamento": ["pensamento"],
                "reclassificacao": ["reclassificacao"],
                "prazo": ["prazo"],
                "compromissoAnalista": ["compromissoAnalista"],
                "necessitaRevisao": ["necessitaRevisao"],
                "resumo_do_andamento": ["resumo_do_andamento"],
            },
            "ai_output_columns": [
                "categoria",
                "pensamento",
                "reclassificacao",
                "prazo",
                "compromissoAnalista",
                "necessitaRevisao",
                "resumo_do_andamento",
            ],
            "input_column_mappings": {
                "numero_processo": ["Numero Processo", "Número Processo"],
                "id_processo": ["ID Processo"],
                "descricao": ["Conteudo", "Conteúdo"],
                "celula": ["Celula", "Célula"],
                "valor_da_causa": ["Valor da Causa"],
                "tipo_de_acao": ["Tipo de Acao", "Tipo de Ação"],
                "marcacao": ["Prazo Agendado"],
                "responsavel": ["Responsavel Publicacao", "Responsável Publicação"],
            },
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = (
        "Número Processo,ID Processo,Conteúdo,Célula,Valor da Causa,Tipo de Ação,Prazo Agendado,Responsável Publicação\n"
        "0001234-56.2026.8.11.0001,42,Descricao importada,C12,50000,Civel,30/03/2026,Equipe A\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    header = [str(item or "") for item in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    assert header == expected_columns
    assert len(header) == 15
    assert row[header.index("numero_processo")] == "0001234-56.2026.8.11.0001"
    assert row[header.index("id_processo")] == "42"
    assert row[header.index("descricao")] == "Descricao importada"
    assert row[header.index("celula")] == "C12"
    assert row[header.index("valor_da_causa")] == "50000"
    assert row[header.index("tipo_de_acao")] == "Civel"
    assert row[header.index("marcacao")] == "30/03/2026"
    assert row[header.index("responsavel")] == "Equipe A"
    assert row[header.index("categoria")] == "Trabalhista"
    assert row[header.index("pensamento")] == "Texto de analise"
    assert row[header.index("reclassificacao")] in {"", None}
    assert row[header.index("prazo")] == "22/03/2026"
    assert row[header.index("compromissoAnalista")] == "Revisar e protocolar"
    assert row[header.index("necessitaRevisao")] == "nao"
    assert row[header.index("resumo_do_andamento")] == "Andamento resumido"


def test_tabular_structured_json_fenced_is_cleaned_and_boolean_normalized(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text=(
                "```json\n"
                "{\n"
                '  "categoria": "\\"ANÁLISE PÓS CITAÇÃO\\",",\n'
                '  "pensamento": "Texto de análise",\n'
                '  "prazo": "\\"Sem prazo\\",",\n'
                '  "compromissoAnalista": true,\n'
                '  "necessitaRevisao": "False",\n'
                '  "Resumo do andamento": "\\"Resumo final\\"}"\n'
                "}\n"
                "```"
            ),
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": [
                "categoria",
                "pensamento",
                "prazo",
                "compromissoAnalista",
                "necessitaRevisao",
                "resumo_do_andamento",
            ],
            "structured_output_aliases": {
                "categoria": ["categoria"],
                "pensamento": ["pensamento"],
                "prazo": ["prazo"],
                "compromissoAnalista": ["compromissoAnalista"],
                "necessitaRevisao": ["necessitaRevisao"],
                "resumo_do_andamento": ["resumo do andamento", "resumo_do_andamento"],
            },
            "ai_output_columns": [
                "categoria",
                "pensamento",
                "prazo",
                "compromissoAnalista",
                "necessitaRevisao",
                "resumo_do_andamento",
            ],
            "include_input_columns": False,
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "Conteudo\nLinha base\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    header = [str(item or "") for item in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    assert row[header.index("categoria")] == "ANÁLISE PÓS CITAÇÃO"
    assert row[header.index("pensamento")] == "Texto de análise"
    assert row[header.index("prazo")] == "Sem prazo"
    assert row[header.index("compromissoAnalista")] == "true"
    assert row[header.index("necessitaRevisao")] == "false"
    assert row[header.index("resumo_do_andamento")] == "Resumo final"


def test_tabular_structured_invalid_json_falls_back_to_textual_parse(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text=(
                '{"categoria":"invalido",\n'
                "categoria: Trabalhista\n"
                "resumo_do_andamento: Andamento por fallback textual\n"
            ),
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": ["categoria", "resumo_do_andamento"],
            "structured_output_aliases": {
                "categoria": ["categoria"],
                "resumo_do_andamento": ["resumo_do_andamento"],
            },
            "ai_output_columns": ["categoria", "resumo_do_andamento"],
            "include_input_columns": False,
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "Conteudo\nLinha base\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    header = [str(item or "") for item in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[header.index("categoria")] == "Trabalhista"
    assert row[header.index("resumo_do_andamento")] == "Andamento por fallback textual"


def test_tabular_input_column_mappings_source_to_target_hydrates_prompt_and_output(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text="resultado: OK",
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Codigo: {{CODIGO}} | Mensagem: {{MENSAGEM}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": ["codigo", "texto", "resultado"],
            "structured_output_aliases": {"resultado": ["resultado"]},
            "ai_output_columns": ["resultado"],
            "input_column_mappings": {
                "Código Fonte": "codigo",
                "Mensagem": "texto",
            },
            "prompt_placeholders": {
                "codigo": "CODIGO",
                "texto": "MENSAGEM",
            },
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "Código Fonte,Mensagem\nABC123,Linha de teste\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 1
    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "ABC123" in sent_prompt
    assert "Linha de teste" in sent_prompt
    assert "{{CODIGO}}" not in sent_prompt
    assert "{{MENSAGEM}}" not in sent_prompt

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    header = [str(item or "") for item in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[header.index("codigo")] == "ABC123"
    assert row[header.index("texto")] == "Linha de teste"
    assert row[header.index("resultado")] == "OK"


def test_tabular_prompt_placeholders_infer_prompt_fields_when_headers_match(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text="resultado: OK",
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Processo: {{NUMERO_PROCESSO}} | Descricao: {{DESCRICAO}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": ["numero_processo", "descricao", "resultado"],
            "structured_output_aliases": {"resultado": ["resultado"]},
            "ai_output_columns": ["resultado"],
            "prompt_placeholders": {
                "numero_processo": "NUMERO_PROCESSO",
                "descricao": "DESCRICAO",
            },
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "Numero Processo,Descricao\n0001234-56.2026.8.11.0001,Linha de teste\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 1
    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "0001234-56.2026.8.11.0001" in sent_prompt
    assert "Linha de teste" in sent_prompt
    assert "{{NUMERO_PROCESSO}}" not in sent_prompt
    assert "{{DESCRICAO}}" not in sent_prompt


def test_tabular_execution_fails_when_placeholder_is_unresolved(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Campo obrigatorio: {{CAMPO_OBRIGATORIO}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": ["documento", "resultado"],
            "structured_output_aliases": {"resultado": ["resultado"]},
            "ai_output_columns": ["resultado"],
            "input_column_mappings": {"Documento": "documento"},
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "Documento\nLinha valida\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "unresolved placeholders" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert provider_service.client.execute_calls == []


def test_text_automation_contract_can_override_output_file_metadata(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(
        analysis_request_id,
        file_name="input.txt",
        mime_type="text/plain",
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        output_type="text_output",
        result_parser="text_raw",
        result_formatter="text_plain",
        output_schema={
            "file_name_template": "saida_{execution_id}.md",
            "mime_type": "text/markdown",
        },
    )
    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "texto de entrada")

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    assert generated["file_name"].startswith("saida_")
    assert generated["file_name"].endswith(".md")
    assert generated["mime_type"] == "text/markdown"


def test_execution_fails_when_explicit_output_contract_is_invalid(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        output_type="invalid_output_type",
    )
    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo")

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "unsupported output_type" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_execution_fails_when_output_schema_payload_is_malformed(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema="{invalid-json",
    )
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: "conteudo\nlinha\n".encode("utf-8"))

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "malformed json payload" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_tabular_prompt_uses_schema_aliases_and_placeholders(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(
            modes=["success"],
            output_text="Classificacao final: OK",
        )
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Documento custom: {{DOCUMENTO_CUSTOM}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "columns": ["linha_origem", "documento", "classificacao_final"],
            "structured_output_aliases": {"classificacao_final": ["classificacao final"]},
            "prompt_field_columns": {"documento": "documento"},
            "prompt_field_aliases": {"documento": ["descricao_custom"]},
            "prompt_placeholders": {"documento": "DOCUMENTO_CUSTOM"},
            "status_column": None,
            "error_column": None,
        },
    )
    csv_payload = "descricao_custom\nTexto de dominio customizado\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 1
    assert "Texto de dominio customizado" in provider_service.client.execute_calls[0]["prompt"]
    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    header = [str(item or "") for item in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    assert "status" not in header
    assert "erro" not in header


def test_tabular_chooses_sheet_with_meaningful_header(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.xlsx")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}"
        ,
        **_legacy_tabular_runtime_kwargs(),
    )

    workbook = Workbook()
    first = workbook.active
    first.title = "vazia"
    first.append([None, None])
    second = workbook.create_sheet("dados")
    second.append(["conteudo"])
    second.append(["linha da aba correta"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: buffer.getvalue())

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert len(provider_service.client.execute_calls) == 1
    assert "linha da aba correta" in provider_service.client.execute_calls[0]["prompt"]


def test_tabular_cost_limit_aborts_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], estimated_cost=Decimal("999.000000"))
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}"
        ,
        **_legacy_tabular_runtime_kwargs(),
    )

    csv_payload = "conteudo\nlinha 1\nlinha 2\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "max_cost_per_execution", 0.10)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value


def test_tabular_primary_with_text_context_applies_context_to_each_row(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    primary_file = _build_request_file(
        analysis_request_id,
        file_name="input.csv",
        file_path="requests/test/primary.csv",
    )
    context_file = _build_request_file(
        analysis_request_id,
        file_name="contexto.pdf",
        file_path="requests/test/contexto.pdf",
        mime_type="application/pdf",
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, primary_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            primary_file.id: primary_file,
            context_file.id: context_file,
        }
    )

    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Conteudo: {{CONTEUDO}}"
        ,
        **_legacy_tabular_runtime_kwargs(),
    )

    csv_payload = (
        "conteudo,prazo agendado\n"
        "Linha 1,2026-01-10\n"
        "Linha 2,2026-01-11\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(
        service,
        "_read_input_file_content",
        lambda **kwargs: "Contexto global de apoio para todas as linhas."
        if kwargs.get("file_path") == context_file.file_path
        else "",
    )

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=primary_file.id,
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=primary_file.id,
            role="primary",
            order_index=0,
        )
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=context_file.id,
            role="context",
            order_index=1,
        )
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 2
    assert all(
        "Contexto global complementar" in call["prompt"]
        and "Contexto global de apoio" in call["prompt"]
        and "[INSTRUCAO]" in call["prompt"]
        and "[DADOS DA LINHA]" in call["prompt"]
        and "[CONTEXTO]" in call["prompt"]
        for call in provider_service.client.execute_calls
    )


def test_multiple_text_inputs_are_combined_for_single_analysis(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    primary_file = _build_request_file(
        analysis_request_id,
        file_name="principal.pdf",
        file_path="requests/test/principal.pdf",
        mime_type="application/pdf",
    )
    context_file = _build_request_file(
        analysis_request_id,
        file_name="apoio.pdf",
        file_path="requests/test/apoio.pdf",
        mime_type="application/pdf",
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, primary_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            primary_file.id: primary_file,
            context_file.id: context_file,
        }
    )

    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    def _read_content(**kwargs):  # type: ignore[no-untyped-def]
        file_path = kwargs.get("file_path")
        if file_path == primary_file.file_path:
            return "Conteudo principal do documento."
        if file_path == context_file.file_path:
            return "Informacoes complementares."
        return ""

    monkeypatch.setattr(service, "_read_input_file_content", _read_content)
    monkeypatch.setattr(execution_module.settings, "chunk_size_characters", 8000)
    monkeypatch.setattr(execution_module.settings, "max_input_characters", 20000)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=primary_file.id,
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=primary_file.id,
            role="primary",
            order_index=0,
        )
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=context_file.id,
            role="context",
            order_index=1,
        )
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 1
    prompt_sent = provider_service.client.execute_calls[0]["prompt"]
    assert "Documento 1 - principal.pdf" in prompt_sent
    assert "Documento 2 - apoio.pdf" in prompt_sent
    assert "Conteudo principal do documento." in prompt_sent
    assert "Informacoes complementares." in prompt_sent

    generated = service.file_service.calls[0]
    assert generated["file_name"].endswith(".txt")
    assert generated["mime_type"] == "text/plain"


def test_multiple_tabular_inputs_are_rejected_with_clear_error() -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    primary_file = _build_request_file(
        analysis_request_id,
        file_name="primary.csv",
        file_path="requests/test/primary.csv",
    )
    extra_tabular = _build_request_file(
        analysis_request_id,
        file_name="context.xlsx",
        file_path="requests/test/context.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, primary_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            primary_file.id: primary_file,
            extra_tabular.id: extra_tabular,
        }
    )
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=primary_file.id,
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=primary_file.id,
            role="primary",
            order_index=0,
        )
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=extra_tabular.id,
            role="context",
            order_index=1,
        )
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "multiple tabular files" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_processing_rejects_inconsistent_roles_without_primary() -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    first_file = _build_request_file(
        analysis_request_id,
        file_name="a.pdf",
        file_path="requests/test/a.pdf",
        mime_type="application/pdf",
    )
    second_file = _build_request_file(
        analysis_request_id,
        file_name="b.pdf",
        file_path="requests/test/b.pdf",
        mime_type="application/pdf",
    )
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, first_file)
    service.request_files = FakeRequestFileRepository(  # type: ignore[assignment]
        {
            first_file.id: first_file,
            second_file.id: second_file,
        }
    )
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=first_file.id,
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=first_file.id,
            role="context",
            order_index=0,
        )
    )
    service.execution_inputs.add(  # type: ignore[attr-defined]
        DjangoAiExecutionInputFile(
            execution_id=execution.id,
            request_file_id=second_file.id,
            role="context",
            order_index=1,
        )
    )

    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "exactly one primary file is required" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_text_prompt_builder_uses_structured_sections() -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)
    execution_profile = service._resolve_execution_profile(automation_id=automation_id)

    prompt = service._build_provider_prompt(
        official_prompt="Resuma o documento de forma objetiva.",
        file_content="Linha 1\n\nLinha 1\nLinha 2",
        execution_profile=execution_profile,
    )

    assert prompt.startswith("[INSTRUCAO]")
    assert "[CONTEXTO]" in prompt
    assert "Arquivo de entrada para analise" in prompt
    assert "Linha 1" in prompt


def test_global_context_prioritizes_type_and_deduplicates_content(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.csv", mime_type="text/csv")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)
    execution_profile = service._resolve_execution_profile(automation_id=automation_id)

    context_inputs = [
        EngineExecutionInput(
            request_file_id=uuid4(),
            role="context",
            order_index=1,
            file_name="ctx_raw.txt",
            file_path="ctx/raw.txt",
            mime_type="text/plain",
            file_kind=ExecutionFileKind.TEXTUAL,
            source="test",
        ),
        EngineExecutionInput(
            request_file_id=uuid4(),
            role="context",
            order_index=1,
            file_name="ctx_structured.json",
            file_path="ctx/structured.json",
            mime_type="application/json",
            file_kind=ExecutionFileKind.TEXTUAL,
            source="test",
        ),
        EngineExecutionInput(
            request_file_id=uuid4(),
            role="context",
            order_index=2,
            file_name="ctx_duplicate.log",
            file_path="ctx/duplicate.log",
            mime_type="text/plain",
            file_kind=ExecutionFileKind.TEXTUAL,
            source="test",
        ),
    ]

    content_map = {
        "ctx/raw.txt": "linha util\n\nlinha util\nlinha final",
        "ctx/structured.json": '{"campo":"valor"}\n{"campo":"valor"}',
        "ctx/duplicate.log": "linha util\nlinha final",
    }
    monkeypatch.setattr(
        service,
        "_read_input_file_content",
        lambda **kwargs: content_map[str(kwargs.get("file_path") or "")],
    )
    monkeypatch.setattr(execution_module.settings, "max_context_file_characters", 5000)
    monkeypatch.setattr(execution_module.settings, "max_context_characters", 5000)

    context_text = service._build_global_context_text(
        context_inputs=context_inputs,
        execution_profile=execution_profile,
    )
    assert context_text is not None
    assert context_text.find("ctx_structured.json") < context_text.find("ctx_raw.txt")
    assert context_text.count("[Contexto ") == 2
    assert "linha util\nlinha util" not in context_text


def test_global_context_applies_per_file_and_total_limits(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.csv", mime_type="text/csv")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)
    execution_profile = service._resolve_execution_profile(automation_id=automation_id)

    context_inputs = [
        EngineExecutionInput(
            request_file_id=uuid4(),
            role="context",
            order_index=0,
            file_name="ctx_a.txt",
            file_path="ctx/a.txt",
            mime_type="text/plain",
            file_kind=ExecutionFileKind.TEXTUAL,
            source="test",
        ),
        EngineExecutionInput(
            request_file_id=uuid4(),
            role="context",
            order_index=1,
            file_name="ctx_b.txt",
            file_path="ctx/b.txt",
            mime_type="text/plain",
            file_kind=ExecutionFileKind.TEXTUAL,
            source="test",
        ),
    ]

    long_text_a = "palavra " * 120
    long_text_b = "diferente " * 120
    monkeypatch.setattr(
        service,
        "_read_input_file_content",
        lambda **kwargs: (
            long_text_a
            if str(kwargs.get("file_path")) == "ctx/a.txt"
            else long_text_b
            if str(kwargs.get("file_path")) == "ctx/b.txt"
            else ""
        ),
    )
    monkeypatch.setattr(execution_module.settings, "max_context_file_characters", 140)
    monkeypatch.setattr(execution_module.settings, "max_context_characters", 260)

    context_text = service._build_global_context_text(
        context_inputs=context_inputs,
        execution_profile=execution_profile,
    )
    assert context_text is not None
    assert len(context_text) <= 260
    assert "contexto truncado para 140 caracteres" in context_text
    assert "contexto truncado para 260 caracteres" in context_text


def test_text_chunks_hard_limit_aborts_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]

    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "A" * 400)
    monkeypatch.setattr(execution_module.settings, "chunk_size_characters", 100)
    monkeypatch.setattr(execution_module.settings, "max_input_characters", 400)
    monkeypatch.setattr(execution_module.settings, "max_text_chunks_hard_limit", 2)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "hard limit of text chunks" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_execution_rows_hard_limit_aborts_tabular_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]

    csv_payload = "conteudo\nlinha 1\nlinha 2\nlinha 3\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 2)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "hard limit of tabular rows" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_provider_calls_hard_limit_aborts_tabular_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]

    csv_payload = "conteudo\nlinha 1\nlinha 2\nlinha 3\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "max_provider_calls_hard_limit", 1)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "hard limit of provider calls" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 1


def test_tabular_row_size_hard_limit_aborts_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]

    csv_payload = "conteudo,campo\nlinha muito grande para o limite hard,valor adicional\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "max_tabular_row_characters_hard_limit", 10)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "row exceeded hard character limit" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_execution_time_hard_limit_aborts_execution(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo basico")
    monkeypatch.setattr(execution_module.settings, "max_execution_seconds_hard_limit", 10)

    timeline = [0.0, 20.0, 20.1, 20.2]

    def _fake_perf_counter() -> float:
        return timeline.pop(0) if timeline else 20.3

    monkeypatch.setattr(execution_module, "perf_counter", _fake_perf_counter)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "hard processing time limit" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_job_retries_hard_limit_marks_execution_failed(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["timeout"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    monkeypatch.setattr(service, "_read_input_file_content", lambda **_: "conteudo basico")
    monkeypatch.setattr(execution_module.settings, "max_retries", 10)
    monkeypatch.setattr(execution_module.settings, "max_job_retries_hard_limit", 1)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
        retry_count=1,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "hard retry limit" in (queue_repo.jobs[queue_job.id].error_message or "").lower()


def test_execution_profile_uses_default_when_no_override(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)

    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(execution_module.settings, "execution_profile_automation_overrides", {})
    monkeypatch.setattr(execution_module.settings, "execution_profile_standard_max_execution_rows", 25000)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 100000)

    resolved = service._resolve_execution_profile(automation_id=automation_id)
    assert resolved.name == "standard"
    assert resolved.source == "env_default"
    assert resolved.max_execution_rows == 25000


def test_execution_profile_can_be_overridden_by_automation(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)

    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(
        execution_module.settings,
        "execution_profile_automation_overrides",
        {str(automation_id).lower(): "heavy"},
    )

    resolved = service._resolve_execution_profile(automation_id=automation_id)
    assert resolved.name == "heavy"
    assert resolved.source == "env_automation_override"


def test_execution_profile_limits_are_clamped_by_hard_limits(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)

    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "extended")
    monkeypatch.setattr(execution_module.settings, "execution_profile_extended_max_execution_rows", 250000)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 100000)

    resolved = service._resolve_execution_profile(automation_id=automation_id)
    assert resolved.max_execution_rows == 100000
    assert "max_execution_rows" in resolved.hard_clamped_fields


def test_profile_rows_limit_can_abort_before_hard_limit(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]

    csv_payload = "conteudo\nlinha 1\nlinha 2\nlinha 3\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(execution_module.settings, "execution_profile_automation_overrides", {})
    monkeypatch.setattr(execution_module.settings, "execution_profile_standard_max_execution_rows", 2)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 100000)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.FAILED.value
    assert "profile limit of tabular rows" in (queue_repo.jobs[queue_job.id].error_message or "").lower()
    assert len(provider_service.client.execute_calls) == 0


def test_automation_override_profile_allows_heavier_tabular_workload(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success", "success", "success"]))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(**_legacy_tabular_runtime_kwargs())  # type: ignore[assignment]

    csv_payload = "conteudo\nlinha 1\nlinha 2\nlinha 3\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)
    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(
        execution_module.settings,
        "execution_profile_automation_overrides",
        {str(automation_id).lower(): "heavy"},
    )
    monkeypatch.setattr(execution_module.settings, "execution_profile_standard_max_execution_rows", 2)
    monkeypatch.setattr(execution_module.settings, "execution_profile_heavy_max_execution_rows", 10)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 100000)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    assert queue_repo.jobs[queue_job.id].job_status == ExecutionStatus.COMPLETED.value
    assert len(provider_service.client.execute_calls) == 3


def test_persisted_profile_has_precedence_over_env_override(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)

    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(
        execution_module.settings,
        "execution_profile_automation_overrides",
        {str(automation_id).lower(): "extended"},
    )
    monkeypatch.setattr(execution_module.settings, "execution_profile_heavy_max_execution_rows", 33333)

    persisted = DjangoAiAutomationExecutionSetting(
        id=uuid4(),
        automation_id=automation_id,
        execution_profile="heavy",
        is_active=True,
    )
    service.execution_profile_settings.active_settings[automation_id] = persisted  # type: ignore[attr-defined]

    resolved = service._resolve_execution_profile(automation_id=automation_id)

    assert resolved.name == "heavy"
    assert resolved.source == "persisted_automation"
    assert resolved.max_execution_rows == 33333
    assert resolved.source_details["origin"] == "persisted_automation"


def test_persisted_profile_override_is_applied_and_hard_clamped(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="entrada.txt", mime_type="text/plain")
    service, _, _ = _build_service(analysis_request_id, automation_id, request_file)

    monkeypatch.setattr(execution_module.settings, "execution_profile_default", "standard")
    monkeypatch.setattr(execution_module.settings, "execution_profile_standard_max_execution_rows", 25000)
    monkeypatch.setattr(execution_module.settings, "max_execution_rows_hard_limit", 100000)

    persisted = DjangoAiAutomationExecutionSetting(
        id=uuid4(),
        automation_id=automation_id,
        execution_profile="standard",
        is_active=True,
        max_execution_rows=120000,
    )
    service.execution_profile_settings.active_settings[automation_id] = persisted  # type: ignore[attr-defined]

    resolved = service._resolve_execution_profile(automation_id=automation_id)

    assert resolved.source == "persisted_automation"
    assert resolved.persisted_overrides["max_execution_rows"] == 120000
    assert resolved.max_execution_rows == 100000
    assert "max_execution_rows" in resolved.hard_clamped_fields


# ---------------------------------------------------------------------------
# Automação tabular: 16 colunas jurídicas com mapeamento de aliases
# ---------------------------------------------------------------------------

_SCHEMA_16_COLUNAS = {
    "columns": [
        "numero_processo",
        "id_processo",
        "id_publicacao",
        "celula",
        "valor_da_causa",
        "tipo_de_acao",
        "marcacao",
        "responsavel",
        "descricao",
        "categoria",
        "pensamento",
        "reclassificacao",
        "prazo",
        "compromissoAnalista",
        "necessitaRevisao",
        "resumo_do_andamento",
    ],
    "structured_output_aliases": {
        "categoria": ["categoria"],
        "pensamento": ["pensamento"],
        "reclassificacao": ["reclassificacao"],
        "prazo": ["prazo"],
        "compromissoAnalista": ["compromissoAnalista"],
        "necessitaRevisao": ["necessitaRevisao"],
        "resumo_do_andamento": ["resumo_do_andamento", "resumo do andamento"],
    },
    "ai_output_columns": [
        "categoria",
        "pensamento",
        "reclassificacao",
        "prazo",
        "compromissoAnalista",
        "necessitaRevisao",
        "resumo_do_andamento",
    ],
    "prompt_placeholders": {
        "numero_processo": "NUMERO_PROCESSO",
        "id_processo": "ID_PROCESSO",
        "id_publicacao": "ID_PUBLICACAO",
        "celula": "CELULA",
        "valor_da_causa": "VALOR_DA_CAUSA",
        "tipo_de_acao": "TIPO_DE_ACAO",
        "marcacao": "MARCACAO",
        "responsavel": "RESPONSAVEL",
        "descricao": "DESCRICAO",
    },
    "prompt_field_columns": {
        "numero_processo": "numero_processo",
        "id_processo": "id_processo",
        "id_publicacao": "id_publicacao",
        "celula": "celula",
        "valor_da_causa": "valor_da_causa",
        "tipo_de_acao": "tipo_de_acao",
        "marcacao": "marcacao",
        "responsavel": "responsavel",
        "descricao": "descricao",
    },
    "input_column_mappings": {
        "numero_processo": ["Número Processo", "Numero Processo", "numero_processo"],
        "id_processo": ["ID Processo", "id_processo"],
        "id_publicacao": ["ID Publicação", "ID Publicacao", "id_publicacao"],
        "celula": ["Célula", "Celula", "celula"],
        "valor_da_causa": ["Valor da Causa", "valor_da_causa"],
        "tipo_de_acao": ["Tipo de Ação", "Tipo de Acao", "tipo_de_acao"],
        "marcacao": ["Marcação", "Marcacao", "Prazo Agendado", "marcacao"],
        "responsavel": ["Responsável Publicação", "Responsavel Publicacao", "responsavel"],
        "descricao": ["Conteúdo", "Conteudo", "Descrição", "Descricao", "descricao"],
    },
    "include_input_columns": False,
    "row_origin_column": None,
    "status_column": None,
    "error_column": None,
}

_PROMPT_JURIDICO = (
    "Numero: {{NUMERO_PROCESSO}} | ID: {{ID_PROCESSO}} | Pub: {{ID_PUBLICACAO}} | "
    "Celula: {{CELULA}} | Valor: {{VALOR_DA_CAUSA}} | Tipo: {{TIPO_DE_ACAO}} | "
    "Marcacao: {{MARCACAO}} | Responsavel: {{RESPONSAVEL}} | Descricao: {{DESCRICAO}}"
)

_AI_RESPONSE_7_COLUNAS = (
    "categoria: Trabalhista\n"
    "pensamento: Análise concluída\n"
    "reclassificacao: Sim\n"
    "prazo: 30 dias\n"
    "compromissoAnalista: Não\n"
    "necessitaRevisao: false\n"
    "resumo_do_andamento: Processo ativo"
)


def test_tabular_coluna_conteudo_acento_hidrata_placeholder_descricao(monkeypatch) -> None:
    """Coluna 'Conteúdo' (com acento) deve hidratar {{DESCRICAO}} via input_column_mappings."""
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], output_text=_AI_RESPONSE_7_COLUNAS)
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=_PROMPT_JURIDICO,
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=_SCHEMA_16_COLUNAS,
    )
    csv_payload = (
        "Numero Processo,ID Processo,ID Publicacao,Celula,Valor da Causa,"
        "Tipo de Acao,Marcacao,Responsavel Publicacao,Conteúdo\n"
        "0001234,P-001,PUB-999,Trabalhista,R$ 10.000,Reclamatória,Urgente,Ana Silva,Fato jurídico relevante\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "Fato jurídico relevante" in sent_prompt
    assert "0001234" in sent_prompt
    assert "{{DESCRICAO}}" not in sent_prompt
    assert "{{NUMERO_PROCESSO}}" not in sent_prompt


def test_tabular_coluna_descricao_sem_acento_hidrata_placeholder_descricao(monkeypatch) -> None:
    """Coluna 'Descricao' (sem acento) deve hidratar {{DESCRICAO}} via alias normalizado."""
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], output_text=_AI_RESPONSE_7_COLUNAS)
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=_PROMPT_JURIDICO,
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=_SCHEMA_16_COLUNAS,
    )
    csv_payload = (
        "Numero Processo,ID Processo,ID Publicacao,Celula,Valor da Causa,"
        "Tipo de Acao,Marcacao,Responsavel Publicacao,Descricao\n"
        "0009999,P-002,PUB-888,Civel,R$ 5.000,Cobrança,Normal,Carlos Melo,Fato descrito sem acento\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "Fato descrito sem acento" in sent_prompt
    assert "{{DESCRICAO}}" not in sent_prompt


def test_tabular_contrato_16_colunas_sem_colunas_legadas(monkeypatch) -> None:
    """Output Excel deve ter exatamente 16 colunas na ordem correta, sem colunas legadas."""
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], output_text=_AI_RESPONSE_7_COLUNAS)
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=_PROMPT_JURIDICO,
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=_SCHEMA_16_COLUNAS,
    )
    csv_payload = (
        "Numero Processo,ID Processo,ID Publicacao,Celula,Valor da Causa,"
        "Tipo de Acao,Marcacao,Responsavel Publicacao,Conteudo\n"
        "0001234,P-001,PUB-999,Trabalhista,R$ 10.000,Reclamatória,Urgente,Ana Silva,Fato de teste\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]

    expected_columns = list(_SCHEMA_16_COLUNAS["columns"])
    assert headers == expected_columns, f"Colunas incorretas: {headers}"
    assert len(headers) == 16

    legacy_columns = {"veredito", "motivo", "classificacao_correta", "trecho_determinante", "resultado"}
    assert not legacy_columns.intersection(set(headers)), f"Coluna legada encontrada: {legacy_columns.intersection(set(headers))}"

    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_dict = dict(zip(headers, row))
    assert str(row_dict.get("numero_processo") or "") == "0001234"
    assert str(row_dict.get("categoria") or "") == "Trabalhista"
    assert str(row_dict.get("resumo_do_andamento") or "") == "Processo ativo"


def test_tabular_campos_entrada_preservados_apos_resposta_ia(monkeypatch) -> None:
    """Os 9 campos de entrada não devem ser sobrescritos pela resposta da IA."""
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], output_text=_AI_RESPONSE_7_COLUNAS)
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=_PROMPT_JURIDICO,
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=_SCHEMA_16_COLUNAS,
    )
    csv_payload = (
        "Numero Processo,ID Processo,ID Publicacao,Celula,Valor da Causa,"
        "Tipo de Acao,Marcacao,Responsavel Publicacao,Conteudo\n"
        "PROC-XYZ,P-777,PUB-123,Trabalhista,R$ 99.000,Rescisão,Alta,Maria Costa,Descrição do fato\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_dict = dict(zip(headers, row))

    assert str(row_dict.get("numero_processo") or "") == "PROC-XYZ", "numero_processo sobrescrito"
    assert str(row_dict.get("id_processo") or "") == "P-777", "id_processo sobrescrito"
    assert str(row_dict.get("descricao") or "") == "Descrição do fato", "descricao sobrescrita"
    assert str(row_dict.get("responsavel") or "") == "Maria Costa", "responsavel sobrescrito"
    assert str(row_dict.get("categoria") or "") == "Trabalhista", "categoria ausente"
    assert str(row_dict.get("resumo_do_andamento") or "") == "Processo ativo", "resumo_do_andamento ausente"


def test_tabular_campos_entrada_preservados_quando_ai_output_columns_incorreto(monkeypatch) -> None:
    """Campos de entrada não devem ser sobrescritos mesmo se ai_output_columns incluir campos de entrada."""
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["success"], output_text=_AI_RESPONSE_7_COLUNAS)
    )
    service.provider_service = provider_service  # type: ignore[assignment]

    schema_ai_cols_errado = {
        **_SCHEMA_16_COLUNAS,
        "ai_output_columns": list(_SCHEMA_16_COLUNAS["columns"]),  # ERRADO: inclui os 9 campos de entrada
    }
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text=_PROMPT_JURIDICO,
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=schema_ai_cols_errado,
    )
    csv_payload = (
        "Numero Processo,ID Processo,ID Publicacao,Celula,Valor da Causa,"
        "Tipo de Acao,Marcacao,Responsavel Publicacao,Conteudo\n"
        "PROC-ABC,P-555,PUB-444,Trabalhista,R$ 50.000,Rescisão,Urgente,João Lima,Fato preservado\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_dict = dict(zip(headers, row))

    assert str(row_dict.get("numero_processo") or "") == "PROC-ABC", "numero_processo sobrescrito por ai_output_columns errado"
    assert str(row_dict.get("descricao") or "") == "Fato preservado", "descricao sobrescrita por ai_output_columns errado"
    assert str(row_dict.get("id_processo") or "") == "P-555", "id_processo sobrescrito"
    assert str(row_dict.get("categoria") or "") == "Trabalhista", "categoria da IA ausente"
    assert str(row_dict.get("resumo_do_andamento") or "") == "Processo ativo", "resumo_do_andamento da IA ausente"


_PURE_JSON_SCHEMA_15_COLUNAS = {
    "type": "object",
    "required": [
        "numero_processo",
        "id_processo",
        "descricao",
        "celula",
        "valor_da_causa",
        "tipo_de_acao",
        "marcacao",
        "responsavel",
        "categoria",
        "pensamento",
        "reclassificacao",
        "prazo",
        "compromissoAnalista",
        "necessitaRevisao",
        "resumo_do_andamento",
    ],
    "properties": {
        "numero_processo": {"type": "string"},
        "id_processo": {"type": "string"},
        "descricao": {"type": "string"},
        "celula": {"type": "string"},
        "valor_da_causa": {"type": "string"},
        "tipo_de_acao": {"type": "string"},
        "marcacao": {"type": "string"},
        "responsavel": {"type": "string"},
        "categoria": {"type": "string"},
        "pensamento": {"type": "string"},
        "reclassificacao": {"type": "string"},
        "prazo": {"type": "string"},
        "compromissoAnalista": {"type": "string"},
        "necessitaRevisao": {"type": "string"},
        "resumo_do_andamento": {"type": "string"},
    },
    "additionalProperties": False,
}


def test_tabular_pure_json_schema_required_infers_columns_and_exact_output(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(
            modes=["success"],
            output_text=(
                "categoria: Trabalhista\n"
                "pensamento: Analise pronta\n"
                "reclassificacao: Analise estrategica\n"
                "prazo: 15\n"
                "compromissoAnalista: Revisar\n"
                "necessitaRevisao: sim\n"
                "resumo_do_andamento: Resumo final"
            ),
        )
    )
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Processo {numero_processo} | Descricao {descricao}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema=_PURE_JSON_SCHEMA_15_COLUNAS,
    )
    csv_payload = (
        "Numero Processo,ID Processo,Conteudo,Celula,Valor da Causa,Tipo de Acao,Marcacao,Responsavel Publicacao\n"
        "000123,ID-9,Descricao original,A1,1000,Civel,Urgente,Equipe A\n"
    ).encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_dict = dict(zip(headers, row))

    assert headers == _PURE_JSON_SCHEMA_15_COLUNAS["required"]
    assert "status" not in headers
    assert "erro" not in headers
    assert "linha_origem" not in headers
    assert str(row_dict["numero_processo"] or "") == "000123"
    assert str(row_dict["descricao"] or "") == "Descricao original"
    assert str(row_dict["categoria"] or "") == "Trabalhista"
    assert str(row_dict["resumo_do_andamento"] or "") == "Resumo final"


def test_tabular_pure_json_schema_properties_order_is_used_when_required_missing(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(  # type: ignore[assignment]
        FakeProviderClient(modes=["success"], output_text='{"resultado":"OK","observacao":"Tudo certo","extra":"ignorar"}')
    )
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Codigo {{codigo}}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "type": "object",
            "properties": {
                "codigo": {"type": "string"},
                "resultado": {"type": "string"},
                "observacao": {"type": "string"},
            },
            "additionalProperties": False,
        },
    )
    csv_payload = "Codigo\nABC-1\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_dict = dict(zip(headers, row))

    assert headers == ["codigo", "resultado", "observacao"]
    assert "extra" not in headers
    assert str(row_dict["codigo"] or "") == "ABC-1"
    assert str(row_dict["resultado"] or "") == "OK"
    assert str(row_dict["observacao"] or "") == "Tudo certo"


def test_tabular_single_brace_placeholder_hydrates_prompt(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(FakeProviderClient(modes=["success"], output_text='{"resultado":"OK"}'))
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Processo {numero_processo} | Descricao {descricao}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "type": "object",
            "required": ["numero_processo", "descricao", "resultado"],
            "properties": {
                "numero_processo": {"type": "string"},
                "descricao": {"type": "string"},
                "resultado": {"type": "string"},
            },
        },
    )
    csv_payload = "Numero Processo,Conteudo\nPROC-1,Descricao hidratada\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    sent_prompt = provider_service.client.execute_calls[0]["prompt"]
    assert "PROC-1" in sent_prompt
    assert "Descricao hidratada" in sent_prompt
    assert "{numero_processo}" not in sent_prompt
    assert "{descricao}" not in sent_prompt


def test_tabular_missing_tabular_schema_fails_without_legacy_fallback(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    service.provider_service = FakeProviderService(FakeProviderClient(modes=["success"]))  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Prompt sem contrato valido",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={"type": "object"},
    )
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: "Conteudo\nLinha\n".encode("utf-8"))

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.FAILED.value
    assert "columns or a JSON Schema with properties" in (queue_repo.jobs[queue_job.id].error_message or "")
    assert len(service.file_service.calls) == 1
    assert service.file_service.calls[0]["file_type"] == "error"


def test_tabular_rate_limit_is_recorded_per_row_without_schema_error(monkeypatch) -> None:
    analysis_request_id = uuid4()
    automation_id = uuid4()
    request_file = _build_request_file(analysis_request_id, file_name="input.csv")
    service, queue_repo, shared_exec_repo = _build_service(analysis_request_id, automation_id, request_file)
    provider_service = FakeProviderService(
        FakeProviderClient(modes=["rate_limit", "success"], output_text='{"resultado":"OK"}')
    )
    service.provider_service = provider_service  # type: ignore[assignment]
    service.runtime_resolver = FakeAutomationRuntimeResolver(  # type: ignore[assignment]
        prompt_text="Codigo {codigo}",
        output_type="spreadsheet_output",
        result_parser="tabular_structured",
        result_formatter="spreadsheet_tabular",
        output_schema={
            "type": "object",
            "required": ["codigo", "resultado", "status_execucao", "erro_execucao"],
            "properties": {
                "codigo": {"type": "string"},
                "resultado": {"type": "string"},
                "status_execucao": {"type": "string"},
                "erro_execucao": {"type": "string"},
            },
            "status_column": "status_execucao",
            "error_column": "erro_execucao",
        },
    )
    csv_payload = "Codigo\nA-1\nA-2\n".encode("utf-8")
    monkeypatch.setattr(service, "_read_input_file_bytes", lambda **_: csv_payload)

    execution, queue_job = _seed_execution_and_job(
        shared_exec_repo=shared_exec_repo,
        queue_repo=queue_repo,
        analysis_request_id=analysis_request_id,
        request_file_id=request_file.id,
    )
    service.process_execution_job(execution_id=execution.id, queue_job_id=queue_job.id, worker_name="worker")

    assert shared_exec_repo.executions[execution.id].status == ExecutionStatus.COMPLETED.value
    generated = service.file_service.calls[0]
    workbook = load_workbook(io.BytesIO(generated["content"]))
    headers = [str(cell or "") for cell in next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    first_row = dict(zip(headers, rows[0]))
    second_row = dict(zip(headers, rows[1]))

    assert str(first_row["codigo"] or "") == "A-1"
    assert str(first_row["status_execucao"] or "") == "erro"
    assert "rate limit" in str(first_row["erro_execucao"] or "").lower()
    assert "schema" not in str(first_row["erro_execucao"] or "").lower()
    assert str(second_row["codigo"] or "") == "A-2"
    assert str(second_row["status_execucao"] or "") == "ok"
    assert str(second_row["resultado"] or "") == "OK"
